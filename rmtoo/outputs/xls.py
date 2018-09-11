'''
 rmtoo
   Free and Open Source Requirements Management Tool

 Output the content to a XLS sheet. Suits love Excel.

 (c) 2018 Kristoffer Nordstrom

 For licensing details see COPYING
'''
from __future__ import unicode_literals

import openpyxl
import datetime

from rmtoo.lib.StdOutputParams import StdOutputParams
from rmtoo.lib.ExecutorTopicContinuum import ExecutorTopicContinuum
from rmtoo.lib.logging import tracer
from rmtoo.lib.CreateMakeDependencies import CreateMakeDependencies


class XlsHandler():
    '''Act as an abstraction layer between rmtoo-objects and OpenPyxl
    related things

    The default configuration is intended to sort the important
    objects. Users can adapt it to their needs.

    The filename is handled in the Xls' parent class.

    '''
    default_config = {
        "output_filename": "artifacts/requirements.xlsx",
        "req_attributes": [
            "ID", "Name", "Topic", "Description", "Status", "Owner",
            "Invented by", "Invented on"],
        "req_sheet": "Requirements",
        "topic_sheet": "Topics"
    }

    def __init__(self, filename, config=None):
        tracer.info("Creating XLS workbook: "
                    + filename)
        self.__filename = filename
        self._cfg = self.default_config
        for key, value in config.items():
            self._cfg[key] = value

        self._wb = openpyxl.Workbook()
        self._ws_req = self._wb.active
        self._ws_req.title = self._cfg['req_sheet']
        self._ws_topics = self._wb.create_sheet(
            self._cfg['topic_sheet'])
        self._ws_cfg = self._wb.create_sheet("Configuration")

        # We require those headers at least
        self._req_headers = self._cfg['req_attributes']
        self._headers = list(self._cfg['req_attributes'])
        self.req_row = None
        self._reqlist = []
        self._topiclist = []

    def write(self):
        self._add_header()
        self._write_reqs()
        self._write_topics()
        self._wb.save(filename=self.__filename)

    def _add_header(self):
        self.req_row = 1
        col = 1
        for val in self._headers:
            self._ws_req.cell(column=col, row=self.req_row, value=val)
            col += 1
        self.req_row += 1

    def _write_reqs(self):
        for req in self._reqlist:
            col = 1
            for key in self._headers:
                if key in req:
                    val = req[key]
                    if isinstance(val, list):
                        val = str(val)
                    self._ws_req.cell(column=col, row=self.req_row,
                                      value=val)
                col += 1
            self.req_row += 1
        self._reqlist = []

    def _write_topics(self):
        row = 1
        for topic in self._topiclist:
            self._ws_topics.cell(column=1, row=row, value=topic.name)
            for key, content in topic.members.items():
                self._ws_topics.cell(column=2, row=row, value=key)
                self._ws_topics.cell(column=3, row=row, value=content)
                row += 1
        self._topiclist = []

    def add_req(self, req):
        req_dict = self._req_extract_dict(req)
        self._check_required_fields(req_dict)
        self._add_new_headers(req_dict)
        self._reqlist.append(req_dict)

    @staticmethod
    def _req_extract_dict(req):
        req_dict = {'ID': req.get_id()}
        for rec in req.record:
            if rec.get_tag() in req.values and isinstance(
                    req.values[rec.get_tag()], datetime.date):
                req_dict[rec.get_tag()] = req.values[rec.get_tag()]
            else:
                req_dict[rec.get_tag()] = rec.get_content()
        return req_dict

    def _check_required_fields(self, req_dict):
        """Will raise KeyError if required fields aren't available"""
        for val in self._req_headers:
            if val not in req_dict.keys():
                tracer.warning("Key (" + val + ") error in " + req_dict['ID'])
            assert val in req_dict.keys()

    def _add_new_headers(self, req_dict):
        for key in list(req_dict.keys()):
            if key not in self._headers:
                self._headers.append(key)

    def add_topic(self, topic):
        class TopicList:
            def __init__(self, name):
                self.name = name
                self.members = {}
        ctop = TopicList(topic.get_topic_name())
        for i in topic.get_tags():
            ctop.members[i.get_tag()] = i.get_content()
        self._topiclist.append(ctop)


class Xls(StdOutputParams, ExecutorTopicContinuum,
          CreateMakeDependencies):

    def __init__(self, oconfig):
        '''Create an openpyxl output object.'''
        tracer.info("Called.")
        StdOutputParams.__init__(self, oconfig)
        CreateMakeDependencies.__init__(self)
        self.__ce3set = None
        self.__fd = None
        self._opiface = XlsHandler(self._output_filename, self._config)

    @staticmethod
    def strescape(string):
        '''Escapes a string: hexifies it.'''
        result = ""
        for fchar in string:
            if ord(fchar) >= 32 and ord(fchar) < 127:
                result += fchar
            else:
                result += "%02x" % ord(fchar)
        return result

    def topic_set_pre(self, _topics_set):
        pass

    def topic_set_post(self, topic_set):
        '''Clean up file.'''
        tracer.debug("Clean up file.")
        self._opiface.write()
        tracer.debug("Finished.")

    def topic_pre(self, topic):
        '''Output one topic.'''
        self._opiface.add_topic(topic)

    def topic_post(self, _topic):
        '''Cleanup things for topic.'''
        pass

    def topic_name(self, name):
        pass

    def topic_text(self, text):
        pass

    def requirement_set_pre(self, rset):
        '''Prepare the requirements set output.'''
        self.__ce3set = rset.get_ce3set()
        self.__testcases = rset.get_testcases()

    def requirement_set_sort(self, list_to_sort):
        '''Sort by id.'''
        return sorted(list_to_sort, key=lambda r: r.get_id())

    def requirement(self, req):
        self._opiface.add_req(req)

    def cmad_topic_continuum_pre(self, _):
        pass