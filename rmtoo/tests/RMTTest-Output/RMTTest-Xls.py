# (c) 2018 Kristoffer Nordstroem, see COPYING

import os
from datetime import date, datetime
import openpyxl
import pytest

from rmtoo.outputs.xls import XlsHandler as xh
from rmtoo.lib.Requirement import Requirement
from rmtoo.tests.lib.Utils import create_tmp_dir, delete_tmp_dir
from rmtoo.tests.lib.TestConfig import TestConfig

LDIR = os.path.dirname(os.path.abspath(__file__))


def create_req(req_id):
    req_txt_items = {}
    inv_date = datetime.strptime('1970-01-01', "%Y-%m-%d").date()
    req_values = {
        'Name': 'Aston Martin DB5',
        'Topic': 'Escape Routes',
        'Description': "Flying out the roof",
        'Priority': 'development',
        'Owner': '007',
        'Invented on': inv_date.isoformat(),
        'Invented by': 'Q',
        'Status': 'not done',
        'Class': 'requirement'
    }
    req = Requirement(u"", req_id, None, None, TestConfig())
    for key, value in req_values.items():
        req.record.append(MockRecordEntry(key, value))
        req_txt_items[key] = MockRecordEntry(None, value)
    req_txt_items['Invented on'] = inv_date
    req.values.update(req_txt_items)
    return req


class MockRecordEntry:
    def __init__(self, tag, txt_str):
        self.__txt_tag = tag
        self.__txt_string = txt_str

    def get_content(self):
        return self.__txt_string

    def get_output_string(self):
        return self.__txt_string

    def get_tag(self):
        return self.__txt_tag


class RMTTestOutputXls:
    "Test-Class for the xlsx output class"

    @classmethod
    def teardown_class(self):
        if self.__tmpdir:
            delete_tmp_dir(self.__tmpdir)

    @classmethod
    def setup_class(self):
        self.__tmpdir = create_tmp_dir()
        self.oconfig = xh.default_config
        self._filename = os.path.join(self.__tmpdir, "reqs.xlsx")
        self.oconfig["output_filename"] = self._filename

    def rmttest_adding_req_header(self):
        self.xlsh = xh(self._filename, self.oconfig)
        self.xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename)
        rws = twb['Requirements']

        i = 0
        for tval in xh.default_config['req_attributes']:
            i += 1
            assert rws.cell(row=1, column=i).value == tval
        assert i == 8
        assert rws['B1'].value == "Name"
        # TODO rm file

    def rmttest_adding_req(self):
        self.xlsh = xh(self._filename, self.oconfig)
        self.xlsh.add_req(create_req(u'SW-101'))
        self.xlsh.add_req(create_req(u'SW-102'))
        self.xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename,
                                     guess_types=True)
        rws = twb['Requirements']
        assert rws['A2'].value == "SW-101"
        assert rws['F2'].value == "007"
        assert rws['H2'].value.date().isoformat() == "1970-01-01"

        assert rws['A3'].value == "SW-102"

    def rmttest_adding_topic(self):
        def mock_hdl():
            pass
        self.xlsh = xh(self._filename, self.oconfig)
        topic_tags = [Mock(**{'get_tag.return_value': "asdf",
                              'get_content.return_value': "qwer"})]
        topic_cfg = {'get_topic_name.return_value': "SuperTopic",
                     'get_tags.return_value': topic_tags}
        topic = Mock(**topic_cfg)
        self.xlsh.add_topic(topic)
        self.xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename,
                                     guess_types=True)
        rws = twb['Topics']
        assert rws['A1'].value == "SuperTopic"
        assert rws['B1'].value == "asdf"
        assert rws['C1'].value == "qwer"
